import time
import numpy as np
from openpyxl import Workbook
from selenium import webdriver,common
import csv


def open_file(path):
    N=1
    i=0
    urls = []
    with open(path, encoding='utf-8') as infile:
        reader = csv.reader(infile)
        for row in reader:
            if i>=N*200:break
            
            if i >= (N-1)*200:
                urls.append(row[0])

            i=i+1
            
        #for row in reader:
        #    urls.append(row[0])
    return urls

#urls = []

def spider_exec(urls):
    #实例化webdriver,executable_path参数指定下载的Chromedriver路径
    driver = webdriver.Chrome(executable_path = r"C:/Users/MSI/Desktop/xxq/chromedriver.exe")
    driver.set_page_load_timeout(10)
    commentors=[]
    books=[]
    star=[]
    times=[]
    allbook=dict()
    i = 1 
    for url in urls:

        page = url

        while(1):
            if i>=4 :break
            #print (i)
            time.sleep(np.random.rand()%5)
            try:
                #打开网页
                driver.get(page)
            except common.exceptions.TimeoutException:
                print('timeout')
                continue
            except:
                print('error')
                print(commentor[-1])
                break
            i = i+1
            try:
                #元素定位
                tags = driver.find_elements_by_css_selector("div.main")
            except:
                break
                
                
            for tag in tags:

                bknm = tag.find_element_by_css_selector("img").get_attribute('title')
                books.append(bknm)
                commentors.append(tag.find_element_by_css_selector("a.name").text)
                times.append(tag.find_element_by_css_selector("span.main-meta").text)
                link = tag.find_element_by_css_selector("a.subject-img").get_attribute('href')

                allbook[bknm]=link
                
                temp=''
                try:
                    temp=tag.find_element_by_css_selector(".main-title-rating").get_attribute('title')
                except:
                    pass
                
                score = 0
                
                if temp == '力荐':
                    score = 5
                elif temp == '推荐':
                    score = 4
                elif temp == '还行':
                    score = 3
                elif temp == '较差':
                    score = 2
                elif temp == '很差':
                    score = 1
                    
                star.append(score)

            try:
                page = driver.find_element_by_css_selector("span.next link").get_attribute('href')
            except:
                print('next user')
                break

    return commentors,star,books,allbook

def save_excel(commentors,star,books,path = 'douban_read.xlsx'):
    wb = Workbook()
    ws = wb.active
    for i in range(int(len(commentors))):
        ws.cell(row = i+1, column = 1).value = commentors[i]
        ws.cell(row = i+1, column = 2).value = star[i]
        ws.cell(row = i+1, column = 3).value = books[i]
    wb.save(path)    

def save_csv(commentors,star,books,path = 'comment_info.csv'):
    with open(path,'a+', encoding='utf-8',newline = '\r\n') as infile:
        writer = csv.writer(infile)
        i=0
        for i in range(len(commentors)):
            writer.writerow([commentors[i],books[i],star[i]])
            
def save_book(allbook,path = 'allbook.csv'):
    
    with open(path, encoding='utf-8') as infile:
        reader = csv.reader(infile)
        for row in reader:
            print(row[1])
            allbook[row[0]]=row[1]
            
    with open(path,'w', encoding='utf-8') as infile:
        #writer = csv.writer(infile)
        for book,link in allbook.items():
            #writer.writerow([book,link])
            infile.write(book+','+link+'\n')

if __name__ =='__main__':

    urls = open_file('alluser.csv')

    commentors,star,books,allbook=spider_exec(urls)
    #commentors,star,books,allbook=spider_exec(['https://www.douban.com/people/frederickhfwang/reviews'])
    #save_excel(commentors,star,books)
    #save_csv(commentors,star,books)
    save_book(allbook)













    
    
