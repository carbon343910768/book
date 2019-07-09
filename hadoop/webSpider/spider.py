import time
import numpy as np
from openpyxl import Workbook
from selenium import webdriver
import csv


def open_file(path):
    urls = []
    with open(path, encoding='utf-8') as infile:
        reader = csv.reader(infile)
        for row in reader:
            urls.append(row[0])
    return urls

#urls = []

def spider_exec(urls):
    #实例化webdriver,executable_path参数指定下载的Chromedriver路径
    driver = webdriver.Chrome(executable_path = r"C:/Users/MSI/Desktop/xxq/chromedriver.exe")
    commentors=[]
    books=[]
    star=[]

    for url in urls:

        page = url

        while(1):
            time.sleep(np.random.rand()*5)
            try:
                
                #打开网页
                driver.get(page)
            except:
                print('error')
                continue
            
            #元素定位
            tags = driver.find_elements_by_css_selector("div.main")
            for tag in tags:
                
                books.append(tag.find_element_by_css_selector("img").get_attribute('title'))
                commentors.append(tag.find_element_by_css_selector("a.name").text)

                temp=''
                try:
                    temp=tag.find_element_by_css_selector(".main-title-rating").get_attribute('title')
                except:
                    pass
                
                score = 0
                
                if temp == '力荐':
                    score = 5
                elif temp == '推荐':
                    score = 4
                elif temp == '还行':
                    score = 3
                elif temp == '较差':
                    score = 2
                elif temp == '很差':
                    score = 1
                    
                star.append(score)

            try:
                page = driver.find_element_by_css_selector("span.next link").get_attribute('href')
            except:
                print('next user')
                break

    return commentors,star,books

def save_excel(commentors,star,books,path = 'douban_read.xlsx'):
    wb = Workbook()
    ws = wb.active
    for i in range(int(len(commentors))):
        ws.cell(row = i+1, column = 1).value = commentors[i]
        ws.cell(row = i+1, column = 2).value = star[i]
        ws.cell(row = i+1, column = 3).value = books[i]
    wb.save(path)    


if __name__ =='__main__':

    urls = open_file('urls.csv')

    commentors,star,books=spider_exec(urls)

    save_excel(commentors,star,books)













    
    
