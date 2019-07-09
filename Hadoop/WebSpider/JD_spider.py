import time
import numpy as np
from openpyxl import Workbook
from selenium import webdriver
import csv


def open_file(path):
    urls = []
    with open(path, encoding='utf-8') as infile:
        reader = csv.reader(infile)
        for row in reader:
            urls.append(row[0])
    return urls

#urls = []

def spider_exec(urls):
    #实例化webdriver,executable_path参数指定下载的Chromedriver路径
    driver = webdriver.Chrome(executable_path = r"C:/Users/MSI/Desktop/xxq/chromedriver.exe")
    #commentors=[]
    #books=[]
    #star=[]
    books = dict()

    for url in urls:

        page = url

        while(1):
            time.sleep(np.random.rand()%5)
            try:
                
                #打开网页
                driver.get(page)
            except:
                print('error')
                continue
            book_imgs = []
            book_name = ''
            img = driver.find_element_by_css_selector('ul.lh').find_elements_by_css_selector('li img')
            for links in img:
                book_imgs.append(links.get_attribute('src'))
                book_name = links.get_attribute('alt')
                #print(links.get_attribute('alt'))
                #print(links.get_attribute('src'))
            books.setdefault(book_name,{})

            details = driver.find_element_by_id('parameter2').text
            #print(detail.text)
            details = details.split('\n')
            for row in details:
                key,value = row.split('：',1)
                books[book_name][key]=value
            try:
                book_price = driver.find_element_by_id('page_maprice').text.split('￥')[1]
            except:
                book_price = driver.find_element_by_id('page_opprice').text.split('￥')[1]
            books[book_name]['价格']=book_price
            

            book_author = driver.find_element_by_css_selector('div.p-author a').get_attribute('data-name')
            books[book_name]['作者']=book_author

            tags = []
            book_tags = driver.find_element_by_id('related-sorts').find_elements_by_css_selector('li')
            for tag in book_tags:
                tags.append(tag.text)

            try:
                #js = "var q=document.getElementsByClassName('main')[0].scrollTop = 10000"  # getElementsByClassName表示获取class='main'的元素列表，0表示第一个，所以使用的时候要加索引
                #driver.execute_script(js)
                #time.sleep(5)
                target = driver.find_element_by_id('detail-tag-id-3')
                driver.execute_script('arguments[0].scrollIntoView();',target)
                book_desc = target.find_element_by_css_selector('div.book-detail-content').text
                #book_desc
            except:
                print('cnm')
                book_desc=''
            books[book_name]['简介']=book_desc
            #作者
            #简介
            #价格

                

            books[book_name]['TAGS']=tags
            books[book_name]['book_images']=book_imgs

            break
    driver.close()
    return books

def save_excel(books,path = 'get_books_200.xlsx'):
    wb = Workbook()
    ws = wb.active
    i = 2
    ws.cell(row = 1, column = 1).value = 'name'
    ws.cell(row = 1, column = 2).value = 'author'
    ws.cell(row = 1, column = 3).value = 'publisher'
    ws.cell(row = 1, column = 4).value = 'descripition'
    ws.cell(row = 1, column = 5).value = 'published'
    ws.cell(row = 1, column = 6).value = 'ISBN'
    ws.cell(row = 1, column = 7).value = 'price'
    for bknm,bkif in books.items():
        ws.cell(row = i, column = 1).value = bknm
        ws.cell(row = i, column = 2).value = bkif['作者']
        ws.cell(row = i, column = 3).value = bkif['出版社']
        ws.cell(row = i, column = 4).value = bkif['简介']
        ws.cell(row = i, column = 5).value = bkif['出版时间']
        ws.cell(row = i, column = 6).value = bkif['ISBN']
        ws.cell(row = i, column = 7).value = bkif['价格']
        i=i+1
    wb.save(path)    

def save_img(books,path='image_urls_200.xlsx'):
    wb = Workbook()
    ws = wb.active
    i = 2
    ws.cell(row = 1, column = 1).value = 'book_name'
    ws.cell(row = 1, column = 2).value = 'urls'

    for bknm,bkif in books.items():
        for url in bkif['book_images']:
            ws.cell(row = i, column = 1).value = bknm
            ws.cell(row = i, column = 2).value = url.replace('/n5/','/n1/')
            i = i+1
    wb.save(path)

def save_tag(books,path='book_tags_200.xlsx'):
    wb = Workbook()
    ws = wb.active
    i = 2
    ws.cell(row = 1, column = 1).value = 'book_name'
    ws.cell(row = 1, column = 2).value = 'tag'

    for bknm,bkif in books.items():
        for tag in bkif['TAGS']:
            ws.cell(row = i, column = 1).value = bknm
            ws.cell(row = i, column = 2).value = tag
            i = i+1
    wb.save(path)


if __name__ =='__main__':

    urls = open_file('cnm1.csv')

    print(urls)
    
    books=spider_exec(urls)

    save_excel(books)

    save_img(books)

    save_tag(books)
